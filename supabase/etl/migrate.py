#!/usr/bin/env python3
"""
QA System V1 — one-time data migration (Google Sheet xlsx → Supabase Postgres).

Stages (run in order, or use --all):
  --identity   AccessControl → app_access/profiles ; Agent_teamLeader_Settings → agents/teams
  --evals      Sheet1 + new_evaluations → qa_evaluations + qa_evaluation_responses
  --workflow   Dispute Cases → qa_disputes ; Coaching Logs → qa_coaching (+ back-fill coached flag)
  --reconcile  Recompute every score from responses and compare to the sheet's stored Score

Usage:
  python migrate.py --all
  python migrate.py --evals --dry-run

Env (see ../../.env.local):
  SUPABASE_DB_URL   postgresql://postgres:[PW]@db.<ref>.supabase.co:5432/postgres
  QA_XLSX_PATH      absolute path to the source workbook

Idempotent: evaluations use the (agent,ticket,eval_date) unique index → re-runs skip dupes.
"""
import argparse, os, sys, datetime as dt
from pathlib import Path

import openpyxl
import psycopg
from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parents[2] / ".env.local")

DB_URL = os.environ.get("SUPABASE_DB_URL")
XLSX   = os.environ.get("QA_XLSX_PATH")
OUT    = Path(__file__).resolve().parent / "_out"
OUT.mkdir(exist_ok=True)

# Sheet1 / new_evaluations: the 22 criteria live in columns index 5..26 (0-based),
# in the same order as the seeded scorecard sort_order 1..22.
CRIT_COL_START = 5
CRIT_COL_COUNT = 22

def norm_email(v):
    return (str(v).strip().lower()) if v else None

def map_result(v):
    s = (str(v).strip().lower()) if v is not None else ""
    if s == "pass": return "pass"
    if s == "fail": return "fail"
    return "na"  # "n/a", blank, anything else

def parse_dt(v):
    if v is None or v == "": return None
    if isinstance(v, (dt.datetime, dt.date)):
        return v if isinstance(v, dt.datetime) else dt.datetime(v.year, v.month, v.day)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try: return dt.datetime.strptime(str(v).strip(), fmt)
        except ValueError: continue
    return None

def map_channel(v):
    s = (str(v).strip().lower()) if v else ""
    if "call" in s: return "Call"
    if "ticket" in s: return "Tickets"
    return "Chat"

def open_ws(name):
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    if name not in wb.sheetnames:
        print(f"  ! sheet '{name}' not found — skipping"); return None, None
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    return header, rows

# ── scorecard (read from DB) ─────────────────────────────────
def load_scorecard(cur):
    cur.execute("select id from qa.qa_scorecards where active limit 1")
    sc = cur.fetchone()
    if not sc: sys.exit("No active scorecard. Apply migration 007 first.")
    sc_id = sc[0]
    cur.execute("""select sort_order, id, weight, is_critical from qa.qa_criteria
                   where scorecard_id=%s and not archived order by sort_order""", (sc_id,))
    by_sort = {r[0]: {"id": r[1], "weight": r[2], "critical": r[3]} for r in cur.fetchall()}
    if len(by_sort) < CRIT_COL_COUNT:
        sys.exit(f"Expected {CRIT_COL_COUNT} criteria, found {len(by_sort)}.")
    return sc_id, by_sort

def score_from(responses, by_sort):
    """Mirror of lib/scoring.ts."""
    score, errs, crit = 100, 0, 0
    for sort_order, result in responses:
        if result != "fail": continue
        c = by_sort.get(sort_order)
        if not c: continue
        errs += 1
        if c["critical"]: crit += 1
        else: score -= c["weight"]
    if crit: score = 0
    return max(score, 0), errs, crit

# ── stages ───────────────────────────────────────────────────
def stage_identity(cur, dry):
    print("» identity: AccessControl, Agent_teamLeader_Settings")
    # users → profiles + app_access(quality)
    header, rows = open_ws("AccessControl")
    n = 0
    if header:
        for r in rows:
            email = norm_email(r[0]); role = (str(r[1]).strip() if len(r) > 1 and r[1] else None)
            if not email or not role: continue
            role = role if role in ("system_owner","super_admin","admin","qa_evaluator","team_lead","agent","viewer") else "viewer"
            if not dry:
                cur.execute("insert into public.profiles(id,email) values (gen_random_uuid(),%s) on conflict (email) do nothing", (email,))
                cur.execute("""insert into public.app_access(email,app,role) values (%s,'quality',%s)
                               on conflict (email,app) do update set role=excluded.role""", (email, role))
            n += 1
    print(f"  users: {n}")
    # agents + team-lead mapping
    header, rows = open_ws("Agent_teamLeader_Settings")
    m = 0
    if header:
        for r in rows:
            tl = norm_email(r[0]); agent = norm_email(r[1]) if len(r) > 1 else None
            if not agent: continue
            if not dry:
                cur.execute("insert into public.profiles(id,email) values (gen_random_uuid(),%s) on conflict (email) do nothing", (agent,))
                cur.execute("""insert into public.agents(email,team_lead_email,active) values (%s,%s,true)
                               on conflict (email) do update set team_lead_email=excluded.team_lead_email""", (agent, tl))
            m += 1
    print(f"  agents: {m}")

def _load_eval_sheet(cur, by_sort, sc_id, sheet, dry, mismatches):
    header, rows = open_ws(sheet)
    if not header: return 0
    loaded = 0
    for r in rows:
        if r is None: continue
        agent = norm_email(r[1]); ticket = (str(r[2]).strip() if r[2] else None)
        if not agent or not ticket: continue
        eval_dt = parse_dt(r[0]) or dt.datetime.utcnow()
        customer = norm_email(r[3]) if len(r) > 3 else None
        channel = map_channel(r[4]) if len(r) > 4 else "Chat"

        responses = []
        for i in range(CRIT_COL_COUNT):
            col = CRIT_COL_START + i
            val = r[col] if len(r) > col else None
            responses.append((i + 1, map_result(val)))  # sort_order = i+1

        score, errs, crit = score_from(responses, by_sort)

        # reconcile against stored Score (col 33 in Sheet1 layout, if present)
        stored = None
        if len(r) > 33 and r[33] not in (None, ""):
            try: stored = round(float(r[33]))
            except (ValueError, TypeError): stored = None
        if stored is not None and stored != score:
            mismatches.append((sheet, ticket, agent, stored, score))

        if dry:
            loaded += 1; continue

        cur.execute("""
            insert into qa.qa_evaluations
              (scorecard_id,agent_email,evaluator_email,ticket_number,customer_email,channel,
               eval_date,score,total_errors,total_critical_errors,status,source,
               notes,areas_for_improvement)
            values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'archived','manual',%s,%s)
            on conflict (agent_email,ticket_number,eval_date) where deleted_at is null
            do nothing
            returning id
        """, (sc_id, agent, agent, ticket, customer, channel, eval_dt, score, errs, crit,
              (str(r[27]) if len(r) > 27 and r[27] else None),
              (str(r[28]) if len(r) > 28 and r[28] else None)))
        got = cur.fetchone()
        if not got: continue   # duplicate skipped
        ev_id = got[0]
        cur.executemany(
            "insert into qa.qa_evaluation_responses(evaluation_id,criterion_id,result) values (%s,%s,%s)",
            [(ev_id, by_sort[so]["id"], res) for so, res in responses])
        loaded += 1
    return loaded

def stage_evals(cur, dry):
    print("» evaluations: Sheet1 + new_evaluations")
    sc_id, by_sort = load_scorecard(cur)
    mismatches = []
    a = _load_eval_sheet(cur, by_sort, sc_id, "Sheet1", dry, mismatches)
    b = _load_eval_sheet(cur, by_sort, sc_id, "new_evaluations", dry, mismatches)
    print(f"  loaded: Sheet1={a}, new_evaluations={b}")
    if mismatches:
        path = OUT / "score_mismatches.csv"
        with open(path, "w", encoding="utf-8") as f:
            f.write("sheet,ticket,agent,sheet_score,computed_score\n")
            for m in mismatches: f.write(",".join(map(str, m)) + "\n")
        total = a + b
        rate = 100 * (1 - len(mismatches) / total) if total else 100
        print(f"  ⚠ score mismatches: {len(mismatches)} ({rate:.2f}% match) → {path}")
    else:
        print("  ✓ all scores reconcile")

def stage_workflow(cur, dry):
    print("» workflow: Dispute Cases, Coaching Logs")
    header, rows = open_ws("Dispute Cases")
    d = 0
    if header:
        for r in rows:
            agent = norm_email(r[3]) if len(r) > 3 else None
            ticket = (str(r[2]).strip() if len(r) > 2 and r[2] else None)
            if not agent: continue
            if not dry:
                cur.execute("""insert into qa.qa_disputes
                    (agent_email,ticket_number,comment,submitted_by,status,
                     evaluation_id)
                    values (%s,%s,%s,%s,'resolved',
                      (select id from qa.qa_evaluations e where e.agent_email=%s and e.ticket_number=%s
                       and e.deleted_at is null order by e.eval_date desc limit 1))""",
                    (agent, ticket, (str(r[4]) if len(r) > 4 and r[4] else None),
                     norm_email(r[5]) if len(r) > 5 else agent, agent, ticket))
            d += 1
    print(f"  disputes: {d}")

    header, rows = open_ws("Coaching Logs")
    c = 0
    if header:
        for r in rows:
            coach = norm_email(r[1]) if len(r) > 1 else None
            agent = norm_email(r[2]) if len(r) > 2 else None
            ticket = (str(r[3]).strip() if len(r) > 3 and r[3] else None)
            if not agent: continue
            ts = parse_dt(r[0])
            if not dry:
                cur.execute("""
                  with ev as (
                    select id from qa.qa_evaluations e
                    where e.agent_email=%s and e.ticket_number=%s and e.deleted_at is null
                    order by e.eval_date desc limit 1
                  )
                  insert into qa.qa_coaching
                    (evaluation_id,agent_email,coach_email,ticket_id,strengths,areas_for_improvement,action_plan,email_sent,created_at)
                  values ((select id from ev),%s,%s,%s,%s,%s,%s,true,%s)
                """, (agent, ticket, agent, coach or agent, ticket,
                      (str(r[4]) if len(r) > 4 and r[4] else None),
                      (str(r[5]) if len(r) > 5 and r[5] else None),
                      (str(r[6]) if len(r) > 6 and r[6] else None),
                      ts or dt.datetime.utcnow()))
                # back-fill coached flag on the matched evaluation
                cur.execute("""update qa.qa_evaluations set coached=true, coached_by=%s, coached_at=%s
                               where agent_email=%s and ticket_number=%s and deleted_at is null""",
                            (coach, ts, agent, ticket))
            c += 1
    print(f"  coaching: {c}")

def stage_reconcile(cur):
    print("» reconcile: recompute scores from DB responses")
    cur.execute("""
      select e.id, e.score,
        100
        - coalesce(sum(case when r.result='fail' and not c.is_critical then c.weight else 0 end),0)
        as base,
        bool_or(r.result='fail' and c.is_critical) as critfail
      from qa.qa_evaluations e
      join qa.qa_evaluation_responses r on r.evaluation_id=e.id
      join qa.qa_criteria c on c.id=r.criterion_id
      where e.deleted_at is null
      group by e.id, e.score
    """)
    bad = 0; total = 0
    for ev_id, stored, base, critfail in cur.fetchall():
        total += 1
        computed = 0 if critfail else max(base, 0)
        if round(float(stored)) != computed: bad += 1
    rate = 100 * (1 - bad / total) if total else 100
    print(f"  checked {total} evaluations · {bad} mismatch · {rate:.2f}% match")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--identity", action="store_true")
    ap.add_argument("--evals", action="store_true")
    ap.add_argument("--workflow", action="store_true")
    ap.add_argument("--reconcile", action="store_true")
    ap.add_argument("--all", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()

    if not DB_URL: sys.exit("Set SUPABASE_DB_URL in .env.local")
    if not XLSX or not Path(XLSX).exists(): sys.exit(f"QA_XLSX_PATH not found: {XLSX}")

    do_all = a.all or not any([a.identity, a.evals, a.workflow, a.reconcile])
    print(f"Connecting to Postgres… (dry-run={a.dry_run})")
    with psycopg.connect(DB_URL, autocommit=False) as conn:
        with conn.cursor() as cur:
            if a.identity or do_all: stage_identity(cur, a.dry_run)
            if a.evals or do_all:    stage_evals(cur, a.dry_run)
            if a.workflow or do_all: stage_workflow(cur, a.dry_run)
            if a.dry_run: conn.rollback(); print("dry-run → rolled back")
            else: conn.commit(); print("committed")
            if a.reconcile or do_all: stage_reconcile(cur)
    print("done.")

if __name__ == "__main__":
    main()
