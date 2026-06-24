#!/usr/bin/env python3
"""
QA data migration over the Supabase REST API (PostgREST) using the service_role key.
Works over IPv4 HTTPS — no direct Postgres connection needed.

Requires the `qa` schema to be exposed (Settings → Data API → Exposed schemas).

Usage:
  python migrate_rest.py --all
  python migrate_rest.py --evals
  python migrate_rest.py --truncate --all     # clear qa data first (fresh reload)

Env (.env.local): NEXT_PUBLIC_SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY, QA_XLSX_PATH,
                  SUPABASE_ACCESS_TOKEN (only needed for --truncate)
"""
import argparse, os, sys, json, urllib.request, datetime as dt
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from supabase import create_client

ROOT = Path(__file__).resolve().parents[2]
load_dotenv(ROOT / ".env.local")

URL   = os.environ["NEXT_PUBLIC_SUPABASE_URL"]
KEY   = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
XLSX  = os.environ.get("QA_XLSX_PATH")
REF   = URL.split("://")[-1].split(".")[0]
OUT   = Path(__file__).resolve().parent / "_out"; OUT.mkdir(exist_ok=True)

CRIT_COL_START = 5
CRIT_COL_COUNT = 22

sb = create_client(URL, KEY)
qa = sb.schema("qa")

# ── helpers ──────────────────────────────────────────────────
def norm_email(v): return (str(v).strip().lower()) if v else None
def clean_ticket(v):
    """Normalize tickets: openpyxl reads numeric cells as floats → '1073377.0'."""
    if v in (None, ""): return None
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit(): s = s[:-2]
    return s
def map_result(v):
    s = (str(v).strip().lower()) if v is not None else ""
    return "pass" if s == "pass" else "fail" if s == "fail" else "na"
def map_channel(v):
    s = (str(v).strip().lower()) if v else ""
    return "Call" if "call" in s else "Tickets" if "ticket" in s else "Chat"
def parse_dt(v):
    if v in (None, ""): return None
    if isinstance(v, (dt.datetime, dt.date)):
        return (v if isinstance(v, dt.datetime) else dt.datetime(v.year, v.month, v.day)).isoformat()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try: return dt.datetime.strptime(str(v).strip(), fmt).isoformat()
        except ValueError: continue
    return None
def chunks(seq, n):
    for i in range(0, len(seq), n): yield seq[i:i+n]
def open_ws(name):
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    if name not in wb.sheetnames:
        print(f"  ! sheet '{name}' not found — skipping"); return None
    ws = wb[name]; it = ws.iter_rows(values_only=True); next(it, None)
    return it

def score_from(responses, by_sort):
    score, errs, crit = 100, 0, 0
    for so, result in responses:
        if result != "fail": continue
        c = by_sort.get(so);  errs += 1
        if not c: continue
        if c["is_critical"]: crit += 1
        else: score -= c["weight"]
    if crit: score = 0
    return max(score, 0), errs, crit

def load_scorecard():
    sc = qa.table("qa_scorecards").select("id").eq("active", True).limit(1).execute().data
    if not sc: sys.exit("No active scorecard. Run setup.sql first.")
    sc_id = sc[0]["id"]
    crit = qa.table("qa_criteria").select("id,sort_order,weight,is_critical").eq("scorecard_id", sc_id).execute().data
    by_sort = {c["sort_order"]: c for c in crit}
    if len(by_sort) < CRIT_COL_COUNT: sys.exit(f"Expected {CRIT_COL_COUNT} criteria, found {len(by_sort)}")
    return sc_id, by_sort

# ── stages ───────────────────────────────────────────────────
def stage_identity():
    print("» identity")
    it = open_ws("AccessControl"); n = 0; rows = []
    if it:
        for r in it:
            email = norm_email(r[0]); role = (str(r[1]).strip() if len(r) > 1 and r[1] else None)
            if not email or not role: continue
            if role not in ("system_owner","super_admin","admin","qa_evaluator","team_lead","agent","viewer"): role = "viewer"
            rows.append({"email": email, "app": "quality", "role": role})
        for c in chunks(rows, 500):
            sb.table("app_access").upsert(c, on_conflict="email,app").execute(); n += len(c)
    print(f"  app_access: {n}")
    it = open_ws("Agent_teamLeader_Settings"); m = 0; arows = {}
    if it:
        for r in it:
            tl = norm_email(r[0]); agent = norm_email(r[1]) if len(r) > 1 else None
            if not agent: continue
            arows[agent] = {"email": agent, "team_lead_email": tl, "active": True}
        for c in chunks(list(arows.values()), 500):
            sb.table("agents").upsert(c, on_conflict="email").execute(); m += len(c)
    print(f"  agents: {m}")

def _load_eval_sheet(sheet, sc_id, by_sort, mismatches):
    it = open_ws(sheet)
    if not it: return 0
    items = []  # (eval_dict, [(sort_order,result)])
    for r in it:
        if r is None: continue
        agent = norm_email(r[1]); ticket = clean_ticket(r[2]) if len(r) > 2 else None
        if not agent or not ticket: continue
        eval_dt = parse_dt(r[0]) or dt.datetime.utcnow().isoformat()
        resp = [(i + 1, map_result(r[CRIT_COL_START + i] if len(r) > CRIT_COL_START + i else None)) for i in range(CRIT_COL_COUNT)]
        score, errs, crit = score_from(resp, by_sort)
        stored = None
        if len(r) > 33 and r[33] not in (None, ""):
            try: stored = round(float(r[33]))
            except (ValueError, TypeError): pass
        if stored is not None and stored != score:
            mismatches.append((sheet, ticket, agent, stored, score))
        items.append(({
            "scorecard_id": sc_id, "agent_email": agent, "evaluator_email": agent,
            "ticket_number": ticket, "customer_email": norm_email(r[3]) if len(r) > 3 else None,
            "channel": map_channel(r[4]) if len(r) > 4 else "Chat", "eval_date": eval_dt,
            "score": score, "total_errors": errs, "total_critical_errors": crit,
            "status": "archived", "source": "manual",
            "notes": (str(r[27]) if len(r) > 27 and r[27] else None),
            "areas_for_improvement": (str(r[28]) if len(r) > 28 and r[28] else None),
        }, resp))
    # de-dupe within sheet on (agent,ticket,eval_date)
    seen = set(); uniq = []
    for ev, resp in items:
        k = (ev["agent_email"], ev["ticket_number"], ev["eval_date"])
        if k in seen: continue
        seen.add(k); uniq.append((ev, resp))
    loaded = 0
    for batch in chunks(uniq, 400):
        ins = qa.table("qa_evaluations").insert([e for e, _ in batch]).execute().data
        resp_rows = []
        for row, (_, resp) in zip(ins, batch):
            for so, res in resp:
                resp_rows.append({"evaluation_id": row["id"], "criterion_id": by_sort[so]["id"], "result": res})
        for rc in chunks(resp_rows, 1000):
            qa.table("qa_evaluation_responses").insert(rc).execute()
        loaded += len(ins)
        print(f"    {sheet}: {loaded}/{len(uniq)}", end="\r")
    print()
    return loaded

def stage_evals():
    print("» evaluations")
    sc_id, by_sort = load_scorecard()
    mismatches = []
    a = _load_eval_sheet("Sheet1", sc_id, by_sort, mismatches)
    b = _load_eval_sheet("new_evaluations", sc_id, by_sort, mismatches)
    total = a + b
    print(f"  loaded {total} (Sheet1={a}, new_evaluations={b})")
    if mismatches:
        p = OUT / "score_mismatches.csv"
        with open(p, "w", encoding="utf-8") as f:
            f.write("sheet,ticket,agent,sheet_score,computed\n")
            for m in mismatches: f.write(",".join(map(str, m)) + "\n")
        rate = 100 * (1 - len(mismatches) / total) if total else 100
        print(f"  ⚠ {len(mismatches)} score mismatches ({rate:.2f}% match) → {p}")
    else:
        print("  ✓ all scores reconcile (100%)")

def _fetch_all_evals():
    """Page past PostgREST's 1000-row cap to build the full lookup map."""
    out = []; start = 0
    while True:
        page = qa.table("qa_evaluations").select("id,agent_email,ticket_number,eval_date").range(start, start + 999).execute().data
        out.extend(page)
        if len(page) < 1000: break
        start += 1000
    return out

def stage_workflow():
    print("» workflow (disputes, coaching)")
    # map (agent, cleaned ticket) → latest eval id for linking
    evs = _fetch_all_evals()
    latest = {}
    for e in evs:
        k = (e["agent_email"], clean_ticket(e["ticket_number"]))
        if k not in latest or e["eval_date"] > latest[k][1]: latest[k] = (e["id"], e["eval_date"])
    it = open_ws("Dispute Cases"); drows = []
    if it:
        for r in it:
            agent = norm_email(r[3]) if len(r) > 3 else None
            ticket = clean_ticket(r[2]) if len(r) > 2 else None
            if not agent: continue
            ev = latest.get((agent, ticket))
            drows.append({"evaluation_id": ev[0] if ev else None, "agent_email": agent,
                          "ticket_number": ticket, "comment": (str(r[4]) if len(r) > 4 and r[4] else None),
                          "submitted_by": norm_email(r[5]) if len(r) > 5 and r[5] else agent, "status": "resolved"})
        for c in chunks(drows, 500): qa.table("qa_disputes").insert(c).execute()
    print(f"  disputes: {len(drows)} ({sum(1 for d in drows if d['evaluation_id'])} linked)")
    it = open_ws("Coaching Logs"); crows = []; coached_keys = []
    if it:
        for r in it:
            coach = norm_email(r[1]) if len(r) > 1 else None
            agent = norm_email(r[2]) if len(r) > 2 else None
            ticket = clean_ticket(r[3]) if len(r) > 3 else None
            if not agent: continue
            ev = latest.get((agent, ticket))
            crows.append({"evaluation_id": ev[0] if ev else None, "agent_email": agent,
                          "coach_email": coach or agent, "ticket_id": ticket,
                          "strengths": (str(r[4]) if len(r) > 4 and r[4] else None),
                          "areas_for_improvement": (str(r[5]) if len(r) > 5 and r[5] else None),
                          "action_plan": (str(r[6]) if len(r) > 6 and r[6] else None), "email_sent": True})
            if ev: coached_keys.append((ev[0], coach, parse_dt(r[0])))
        for c in chunks(crows, 500): qa.table("qa_coaching").insert(c).execute()
        # back-fill coached flag
        for ev_id, coach, ts in coached_keys:
            qa.table("qa_evaluations").update({"coached": True, "coached_by": coach, "coached_at": ts}).eq("id", ev_id).execute()
    print(f"  coaching: {len(crows)}")

def truncate():
    tok = os.environ.get("SUPABASE_ACCESS_TOKEN")
    if not tok: sys.exit("--truncate needs SUPABASE_ACCESS_TOKEN")
    sql = "truncate qa.qa_evaluation_responses, qa.qa_evaluations, qa.qa_disputes, qa.qa_coaching restart identity cascade;"
    req = urllib.request.Request(f"https://api.supabase.com/v1/projects/{REF}/database/query",
        data=json.dumps({"query": sql}).encode(),
        headers={"Authorization": f"Bearer {tok}", "Content-Type": "application/json",
                 "User-Agent": "Mozilla/5.0 Chrome/124"})
    urllib.request.urlopen(req, timeout=60); print("» truncated qa data tables")

def main():
    ap = argparse.ArgumentParser()
    for f in ("identity", "evals", "workflow", "all", "truncate"): ap.add_argument(f"--{f}", action="store_true")
    a = ap.parse_args()
    if not XLSX or not Path(XLSX).exists(): sys.exit(f"QA_XLSX_PATH not found: {XLSX}")
    do_all = a.all or not any([a.identity, a.evals, a.workflow])
    if a.truncate: truncate()
    if a.identity or do_all: stage_identity()
    if a.evals or do_all:    stage_evals()
    if a.workflow or do_all: stage_workflow()
    print("done.")

if __name__ == "__main__":
    main()
